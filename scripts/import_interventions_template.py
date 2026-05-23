import json
import time
import urllib.request
from pathlib import Path

import openpyxl


WORKBOOK = Path(r"C:\Users\Jeff Factora\Downloads\interventions_template_1.xlsx")
ENV_FILE = Path("app/.env.production")
FISCAL_YEAR = "2027"
SOURCE_FILENAME = WORKBOOK.name


def read_endpoint():
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        if line.startswith("VITE_APPS_SCRIPT_URL="):
            return line.split("=", 1)[1].strip()
    raise RuntimeError("VITE_APPS_SCRIPT_URL not found in app/.env.production")


def clean(value):
    return "" if value is None else str(value).strip()


def numeric(value):
    if value in (None, ""):
        return 0
    return float(value)


def load_rows():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    records = []
    interventions = {}
    source_row = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        source_row += 1
        record = dict(zip(headers, row))
        if not any(value not in (None, "") for value in record.values()):
            continue

        intervention = clean(record.get("Intervention"))
        municipality = clean(record.get("Municipality"))
        province = clean(record.get("Province"))
        office = clean(record.get("Implementing Unit"))
        commodity = clean(record.get("Commodity"))
        if intervention:
            interventions.setdefault(intervention.lower(), intervention)

        for tier_label, column in (("Tier 1", "Budget Tier 1"), ("Tier 2", "Budget Tier 2")):
            amount = numeric(record.get(column))
            if not amount:
                continue
            records.append({
                "id": f"IMP-2027-{source_row:05d}-{tier_label.replace(' ', '').upper()}",
                "fiscal_year": FISCAL_YEAR,
                "title": " - ".join(part for part in [intervention, commodity, municipality, tier_label] if part),
                "description": "",
                "office": office,
                "program": "",
                "subprogram": "",
                "mfo": "",
                "pap": "",
                "uacs": "",
                "province": province,
                "municipality": municipality,
                "district": "",
                "commodity": commodity,
                "intervention_type": intervention,
                "beneficiary_group": "",
                "beneficiaries": "",
                "budget_amount": amount,
                "nep_amount": "",
                "gaa_amount": "",
                "tier": tier_label,
                "source": "",
                "justification": "",
                "expected_output": "",
                "expected_outcome": "",
                "readiness_status": "",
                "climate_tag": "",
                "climate_rationale": "",
                "gedsi_tag": "",
                "schedule": "",
                "remarks": "",
                "validation_status": "Needs Correction",
                "current_phase": "Proposal",
                "created_at": "",
                "updated_at": "",
                "created_by": "Bulk import",
                "updated_by": "Bulk import",
            })
    workbook.close()
    return records, sorted(interventions.values(), key=str.lower)


def intervention_master_rows(interventions):
    now = "2026-05-23T00:00:00+08:00"
    rows = []
    for index, intervention in enumerate(interventions, 1):
        rows.append({
            "id": f"XLSX-INT-{index:04d}",
            "name": intervention,
            "program": "",
            "mfo": "",
            "source_indicator": "",
            "source_file": SOURCE_FILENAME,
            "created_at": now,
            "updated_at": now,
            "created_by": "Bulk import",
            "updated_by": "Bulk import",
        })
    return rows


def post(endpoint, action, payload):
    body = json.dumps({"action": action, "payload": payload}).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={"Content-Type": "text/plain;charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        data = json.loads(response.read().decode("utf-8"))
    if not data.get("ok"):
        raise RuntimeError(data.get("error") or data)
    return data.get("data")


def post_with_retry(endpoint, action, payload, attempts=4):
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return post(endpoint, action, payload)
        except Exception as error:
            last_error = error
            wait = min(20, attempt * 3)
            print(f"request failed on attempt {attempt}/{attempts}: {error}; retrying in {wait}s")
            time.sleep(wait)
    raise last_error


def chunks(values, size):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def main():
    endpoint = read_endpoint()
    records, interventions = load_rows()
    summary = {
        "source": str(WORKBOOK),
        "proposal_rows_to_upsert": len(records),
        "unique_interventions": len(interventions),
        "first_interventions": interventions[:10],
    }
    Path("C:/tmp/interventions_import_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    Path("C:/tmp/interventions_unique.txt").write_text("\n".join(interventions), encoding="utf-8")

    intervention_rows = intervention_master_rows(interventions)
    for index, batch in enumerate(chunks(intervention_rows, 200), 1):
        result = post_with_retry(endpoint, "upsertRows", {
            "sheetName": "intervention_types",
            "keyField": "id",
            "rows": batch,
        })
        print(f"intervention batch {index}: {result}")

    for index, batch in enumerate(chunks(records, 500), 1):
        result = post_with_retry(endpoint, "upsertRows", {
            "sheetName": "proposals",
            "keyField": "id",
            "rows": batch,
        })
        print(f"proposal batch {index}: {result}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
