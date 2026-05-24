import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def number(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0


def text(value):
    return str(value or "").strip()


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: extract-spatial-workbook.py <workbook.xlsx>")
    path = Path(sys.argv[1])
    wb = load_workbook(path, read_only=True, data_only=True)
    entries = []
    commodities = set()
    offices = set()
    municipalities = set()

    if "BY INTERVENTION" in wb.sheetnames:
        ws = wb["BY INTERVENTION"]
        for row_index in range(10, ws.max_row + 1):
            region = text(ws.cell(row_index, 1).value)
            province = text(ws.cell(row_index, 2).value)
            municipality = text(ws.cell(row_index, 3).value)
            office = text(ws.cell(row_index, 10).value)
            commodity = text(ws.cell(row_index, 27).value)
            if municipality and municipality.lower() != "[not specified]":
                municipalities.add((municipality, province))
            if office:
                offices.add(office)
            if commodity:
                commodities.add(commodity)

            for source_label, intervention_col, tier1_col, tier2_col, suffix in [
                ("LGU-identified intervention", 4, 5, 6, "A"),
                ("Additional DA OU intervention", 7, 8, 9, "B"),
            ]:
                intervention = text(ws.cell(row_index, intervention_col).value)
                tier1 = number(ws.cell(row_index, tier1_col).value)
                tier2 = number(ws.cell(row_index, tier2_col).value)
                if not intervention or not (tier1 or tier2):
                    continue
                entries.append({
                    "sheet": ws.title,
                    "row": row_index,
                    "suffix": suffix,
                    "sourceLabel": source_label,
                    "region": region,
                    "province": province,
                    "municipality": municipality,
                    "office": office,
                    "commodity": commodity,
                    "intervention": intervention,
                    "tier1Thousands": tier1,
                    "tier2Thousands": tier2,
                })

    if "BY COMMODITY or INDUSTRY" in wb.sheetnames:
        ws = wb["BY COMMODITY or INDUSTRY"]
        for row_index in range(10, ws.max_row + 1):
            province = text(ws.cell(row_index, 2).value)
            municipality = text(ws.cell(row_index, 3).value)
            commodity = text(ws.cell(row_index, 4).value) or text(ws.cell(row_index, 7).value)
            office = text(ws.cell(row_index, 10).value)
            if municipality and municipality.lower() != "[not specified]":
                municipalities.add((municipality, province))
            if commodity:
                commodities.add(commodity)
            if office:
                offices.add(office)

    print(json.dumps({
        "sourceFile": str(path),
        "entries": entries,
        "commodities": sorted(commodities),
        "offices": sorted(offices),
        "municipalities": [{"name": name, "province": province, "district": ""} for name, province in sorted(municipalities)],
    }))


if __name__ == "__main__":
    main()
